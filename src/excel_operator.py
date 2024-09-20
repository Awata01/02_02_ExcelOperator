import os
from typing import List
import openpyxl
import csv
import win32com.client
import re
class ExcelOperator:

    @staticmethod
    def get_files_in_path(search_path: str, sub_dir_flg: bool = False, xlsm_flg: bool = False) -> List[str]:
        """
        指定したパス配下の全てのExcelファイルを取得

        :param search_path  :検索するディレクトリのパス
        :param sub_dir_flg  :True   サブディレクトリ配下を含めて走査
                            False   指定パス直下で走査(デフォルト)
        :param xlsmFlg      :True   xlsm ファイルも取得
                            False   xlsx ファイルのみ取得（デフォルト）
        :return             :発見したExcelファイルパスのリスト
        """
        excel_files = []
        # 対象とする拡張子を設定
        valid_extensions = ['.xlsx', '.xlsm'] if xlsm_flg else ['.xlsx']

        # サブディレクトリを含めて検索するかを選択
        search_method = os.walk(search_path) if sub_dir_flg else [(search_path, [], os.listdir(search_path))]

        # ファイルをフィルタリングしてリストに格納(一時ファイルは除外)
        excel_files = [
            os.path.join(root, f) for root, _, files in search_method
            for f in files
            if any(f.endswith(ext) for ext in valid_extensions) and not f.startswith('~$')
        ]
        return excel_files

    @staticmethod
    def get_sheets_name(excel_file_path: str) -> List[str]:
        """
        Excelファイルの全てのシート名を取得

        :param file_path    :対象のExcelファイルパス
        :return             :発見したシート名のリスト
        """
        wb = openpyxl.load_workbook(excel_file_path)
        return wb.sheetnames

    @staticmethod
    def sort_sheet(excel_file_path: str, order: str = 'asc') -> List[str]:
        """
        指定ブックのシートを名前の昇順/降順に並び替え

        :param file_path    :対象のExcelファイルパス
        :param order        :asc    シート名の昇順で並び替え（デフォルト）
                            desc    シート名の降順で並び替え
        :return             :発見したシート名のリスト
        """
        # order引数の正値確認
        order = order.lower()
        if order not in ['asc', 'desc']:
            raise ValueError("Invalid value for 'order'. Must be 'asc' or 'desc'.")

        wb = openpyxl.load_workbook(excel_file_path)
        sorted_sheets = sorted(wb.sheetnames, reverse=(order == 'desc'))
        return sorted_sheets

    @staticmethod
    def convert_csv(excel_file_path: str, sheet_name: str, csv_file_path: str) -> None:
        """
        指定ブックの指定シートをcsvファイル化

        :param excel_file_path  :対象のExcelファイルパス
        :param sheet_name       :CSV化する対象のシート名
        :param csv_file_path    :出力先のCSVファイルパス
        :return                 :None
        """
        wb = openpyxl.load_workbook(excel_file_path)
        # 指定されたシート名が存在するか確認
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the Excel file.")

        # シートを取得
        ws = wb[sheet_name]

        # CSVファイルに書き込む
        with open(csv_file_path, mode='w', newline='', encoding='utf-8-sig') as csvfile:
            csv_writer = csv.writer(csvfile)
            # セル値のみ書き込み
            for row in ws.iter_rows(values_only=True):
                csv_writer.writerow(row)

    @staticmethod
    def change_font(excel_file_path: str, sheet_name: str, font_name: str = 'Meiryo UI'):
        """
        指定シートの全テキストを任意のフォントに変更

        :param excel_file_path  :対象のExcelファイルパス
        :param sheet_name       :フォントを変更する対象のシート名
        :param font_name        :変更後のフォント名（デフォルトで'Meiryo UI'）
        :return                 :None
        """
        # Excelファイルを読み込む
        wb = openpyxl.load_workbook(excel_file_path)
        # 指定したシートを取得
        ws = wb[sheet_name]
        # フォントを作成
        custom_font = openpyxl.styles.Font(name=font_name)

        # シート内の全ての入力済みセルにフォントを適用
        for row in ws.iter_rows():
            for cell in row:
                cell.font = custom_font

        # 変更を保存
        wb.save(excel_file_path)

    @staticmethod
    def search_string_in_book(
            excel_file_path: str,
            search_string: str,
            exact_match: bool = False,
            use_regex: bool = False) -> list:
        """
        Excelワークブック内で文字列を検索し、シート名、セルの位置、そして該当する文字列を返す。
        図形内の文字列も検索対象に含める。

        :param file_name: 検索対象のExcelファイルの絶対パス
        :param search_string: 検索する文字列
        :param exact_match: Trueの場合、完全一致検索を行う。Falseの場合、部分一致検索を行う。
        :param use_regex: Trueの場合、正規表現を使用して検索する。Falseの場合、文字列検索を行う。
        :return: 検索結果のリスト。各要素はタプル (シート名, 'A1' などのセルの位置, セルの値) 形式。
        """
        # 指定されたファイル名からフルパスを作成

        # Excelアプリケーションを起動
        excel = win32com.client.Dispatch("Excel.Application")

        # Excelファイルを開く
        wb = excel.Workbooks.Open(excel_file_path, ReadOnly=True)

        # 検索結果を格納するリスト
        results = []

        # 正規表現パターンを作成
        if use_regex:
            if exact_match:
                pattern = re.compile(f"^{re.escape(search_string)}$")
            else:
                pattern = re.compile(search_string)
        else:
            pattern = None

        # ワークブック内の全てのシートをループ処理
        for sheet in wb.Sheets:
            sheet_name = sheet.Name
            # セルの検索
            used_range = sheet.UsedRange
            for row in used_range.Rows:
                for cell in row.Cells:
                    cell_value = str(cell.Value) if cell.Value is not None else ""
                    if use_regex:
                        if pattern.search(cell_value):
                            col_letter = chr(cell.Column + 64)
                            results.append((sheet_name, f"{col_letter}{cell.Row}", cell_value))
                    else:
                        if exact_match:
                            if cell_value == search_string:
                                col_letter = chr(cell.Column + 64)
                                results.append((sheet_name, f"{col_letter}{cell.Row}", cell_value))
                        else:
                            if search_string in cell_value:
                                col_letter = chr(cell.Column + 64)
                                results.append((sheet_name, f"{col_letter}{cell.Row}", cell_value))

            # 図形内のテキストボックスを検索
            for shape in sheet.Shapes:
                shape_text = shape.TextFrame2.TextRange.Text
                shape_name = shape.Name  # 図形の名前を取得
                if use_regex:
                    if pattern.search(shape_text):
                        results.append((sheet_name, f"図形: {shape_name}", shape_text))
                else:
                    if exact_match:
                        if shape_text == search_string:
                            results.append((sheet_name, f"図形: {shape_name}", shape_text))
                    else:
                        if search_string in shape_text:
                            results.append((sheet_name, f"図形: {shape_name}", shape_text))

        # Excelファイルを閉じる
        wb.Close(False)
        excel.Quit()

        # 検索結果のリストを返す
        return results

    @staticmethod
    def replace_string_in_book(
            excel_file_path: str,
            search_string: str,
            replace_string: str,
            exact_match: bool = False,
            use_regex: bool = False) -> list:
        """
        Excelワークブック内で文字列を検索し、指定された文字列で置換する。
        置換が行われたシート名、セルの位置、そして置換後の値を返す。

        :param excel_file_path: 検索対象のExcelファイルの絶対パス
        :param search_string: 検索する文字列
        :param replace_string: 置換後の文字列
        :param exact_match: Trueの場合、完全一致で置換。Falseの場合、部分一致で置換。
        :param use_regex: Trueの場合、正規表現を使用して置換する。Falseの場合、文字列置換を行う。
        :return: 置換結果のリスト。各要素はタプル (シート名, 'A1' などのセルの位置, 置換前の値, 置換後の値) 形式。
        """
        # 指定されたファイル名からフルパスを作成
        print(f"Opening file: {excel_file_path}")  # デバッグ用: ファイルパスを表示

        # Excelアプリケーションを起動
        excel = win32com.client.Dispatch("Excel.Application")

        # Excelファイルを開く
        try:
            wb = excel.Workbooks.Open(excel_file_path, ReadOnly=False)
        except Exception as e:
            print(f"Failed to open file: {e}")
            return []
        # 置換結果を格納するリスト
        results = []

        # 正規表現パターンを作成
        if use_regex:
            if exact_match:
                pattern = re.compile(f"^{re.escape(search_string)}$")
            else:
                pattern = re.compile(search_string)
        else:
            pattern = None

        # ワークブック内の全てのシートをループ処理
        for sheet in wb.Sheets:
            sheet_name = sheet.Name
            # セルの置換
            used_range = sheet.UsedRange
            for row in used_range.Rows:
                for cell in row.Cells:
                    cell_value = str(cell.Value) if cell.Value is not None else ""
                    new_value = cell_value

                    if use_regex:
                        if pattern.search(cell_value):
                            new_value = pattern.sub(replace_string, cell_value)
                    else:
                        if exact_match:
                            if cell_value == search_string:
                                new_value = replace_string
                        else:
                            if search_string in cell_value:
                                new_value = cell_value.replace(search_string, replace_string)

                    # 値が変更された場合、結果を追加
                    if new_value != cell_value:
                        col_letter = chr(cell.Column + 64)
                        cell.Value = new_value  # Excel内で値を置換
                        results.append((sheet_name, f"{col_letter}{cell.Row}", cell_value, new_value))

            # 図形内のテキストボックスを置換
            for shape in sheet.Shapes:
                shape_text = shape.TextFrame2.TextRange.Text
                new_shape_text = shape_text

                if use_regex:
                    if pattern.search(shape_text):
                        new_shape_text = pattern.sub(replace_string, shape_text)
                else:
                    if exact_match:
                        if shape_text == search_string:
                            new_shape_text = replace_string
                    else:
                        if search_string in shape_text:
                            new_shape_text = shape_text.replace(search_string, replace_string)

                # 値が変更された場合、結果を追加
                if new_shape_text != shape_text:
                    shape.TextFrame2.TextRange.Text = new_shape_text  # 図形内の値を置換
                    results.append((sheet_name, f"図形: {shape.Name}", shape_text, new_shape_text))

        # Excelファイルを保存し閉じる
        wb.Save()
        wb.Close(False)
        excel.Quit()
        # 置換結果のリストを返す
        return results

    @staticmethod
    def set_grid_size(excel_file_path: str, sheet_name: str, pixel_size: int) -> None:
        """
        指定ブックの指定シートのセルサイズを方眼紙サイズに設定

        :param excel_file_path: 対象のExcelファイルの絶対パス
        :param sheet_name: 対象のシート名
        :param pixel_size: 方眼紙のセルサイズ（ピクセル）
        :return: なし
        """
        # Excelアプリケーションを起動
        excel = win32com.client.Dispatch("Excel.Application")

        # Excelファイルを開く
        try:
            wb = excel.Workbooks.Open(excel_file_path, ReadOnly=False)
        except Exception as e:
            print(f"ファイルを開くのに失敗しました: {e}")
            return []

        # 指定されたシートを取得
        ws = wb.Worksheets(sheet_name)

        # ピクセルから行高さと列幅をポイントに変換
        row_size = pixel_size * 0.75  # 行の高さはおおよそピクセルの0.75倍で計算
        col_size = pixel_size * 0.14  # 列の幅はピクセルの0.14倍程度で調整

        # 全てのセルに行の高さと列の幅を適用
        ws.Cells.RowHeight = row_size
        ws.Cells.ColumnWidth = col_size

        # 保存する
        try:
            # Excelファイルを保存して閉じる
            wb.Close(True)
            print(f"{sheet_name} シートのセルサイズを {pixel_size} ピクセルに設定しました。")
        except Exception as e:
            print(f"ファイルの保存に失敗しました: {e}")